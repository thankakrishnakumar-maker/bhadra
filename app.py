import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import uuid
import base64
import time
import json
import io
import urllib.parse
import csv

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="🛕 Sree Bhadreshwari Amman Temple",
    page_icon="🛕",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# SUPABASE CONNECTION
# ============================================================
try:
    from supabase import create_client, Client
    SUPABASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

    @st.cache_resource
    def get_supabase_client():
        return create_client(SUPABASE_URL, SUPABASE_KEY)

    supabase: Client = get_supabase_client()
    DB_CONNECTED = True
except Exception as e:
    DB_CONNECTED = False
    st.error(f"Database connection failed: {str(e)}")

# ============================================================
# BARCODE GENERATION
# ============================================================
BARCODE_AVAILABLE = False
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    try:
        pass
    except:
        pass

# ============================================================
# QR CODE GENERATION (fallback for barcodes)
# ============================================================
QRCODE_AVAILABLE = False
try:
    import qrcode
    QRCODE_AVAILABLE = True
except ImportError:
    pass

# ============================================================
# PDF GENERATION
# ============================================================
PDF_AVAILABLE = False
try:
    from fpdf import FPDF

    class BillPDF(FPDF):
        def __init__(self, amman_img_path=None):
            super().__init__()
            self.amman_img_path = amman_img_path

        def header(self):
            # Add Amman logo if available
            if self.amman_img_path:
                try:
                    self.image(self.amman_img_path, 10, 8, 25)
                    self.image(self.amman_img_path, 175, 8, 25)
                except Exception:
                    pass

            self.set_font('Helvetica', 'B', 16)
            self.cell(0, 10, 'Sree Bhadreshwari Amman Temple', 0, 1, 'C')
            self.set_font('Helvetica', '', 10)
            self.cell(0, 6, 'Amme Narayana .. Devi Narayana', 0, 1, 'C')
            self.set_font('Helvetica', '', 8)
            self.cell(0, 5, 'Temple Management System - Official Receipt', 0, 1, 'C')
            self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
            self.ln(5)

        def footer(self):
            self.set_y(-30)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(3)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 5, 'Thank you for your contribution! May Goddess bless you!', 0, 1, 'C')
            self.cell(0, 5, 'Amme Narayana .. Devi Narayana', 0, 1, 'C')
            self.cell(0, 5, f'Generated on: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}', 0, 1, 'C')

    def save_base64_image_to_temp(base64_str):
        """Save base64 image to temp file for PDF usage"""
        if not base64_str:
            return None
        try:
            import tempfile
            if ',' in base64_str:
                header, data = base64_str.split(',', 1)
            else:
                data = base64_str

            img_data = base64.b64decode(data)

            # Determine extension
            ext = '.png'
            if base64_str.startswith('data:image/jpeg') or base64_str.startswith('data:image/jpg'):
                ext = '.jpg'

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            tmp.write(img_data)
            tmp.close()
            return tmp.name
        except Exception:
            return None

    def generate_bill_pdf(bill_no, manual_bill, bill_book, bill_date,
                          name, address, mobile, pooja_type, amount,
                          amman_base64=None):
        # Try to get amman image for PDF
        amman_path = None
        if amman_base64 and not amman_base64.startswith('data:image/svg'):
            amman_path = save_base64_image_to_temp(amman_base64)

        pdf = BillPDF(amman_img_path=amman_path)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=35)

        # Bill details box
        pdf.set_font('Helvetica', '', 10)
        pdf.set_fill_color(255, 248, 240)
        y_start = pdf.get_y()
        pdf.rect(10, y_start, 190, 100, 'D')

        pdf.set_xy(15, y_start + 5)
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(0, 8, 'BILL / RECEIPT', 0, 1, 'C')
        pdf.ln(2)

        # Bill info
        for label, value in [
            ("Bill No", str(bill_no or '')),
            ("Manual Bill No", str(manual_bill or '')),
            ("Bill Book No", str(bill_book or '')),
            ("Date", str(bill_date or ''))
        ]:
            pdf.set_x(15)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(50, 7, f"{label}:", 0, 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, value, 0, 1)

        pdf.ln(2)
        pdf.line(15, pdf.get_y(), 195, pdf.get_y())
        pdf.ln(3)

        # Devotee info
        for label, value in [
            ("Name", str(name or '')),
            ("Address", str(address or '')),
            ("Mobile", str(mobile or ''))
        ]:
            pdf.set_x(15)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(50, 7, f"{label}:", 0, 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, value, 0, 1)

        pdf.ln(5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)

        # Pooja & Amount
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_x(15)
        pdf.cell(50, 8, "Pooja Type:", 0, 0)
        pdf.set_font('Helvetica', '', 12)
        pdf.cell(0, 8, str(pooja_type or ''), 0, 1)

        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_x(15)
        pdf.cell(50, 12, "Amount:", 0, 0)
        pdf.set_text_color(0, 128, 0)
        pdf.cell(0, 12, f"Rs. {float(amount):,.2f}", 0, 1)
        pdf.set_text_color(0, 0, 0)

        pdf.ln(5)

        # Amount in words (simple)
        pdf.set_font('Helvetica', 'I', 9)
        pdf.set_x(15)
        pdf.cell(0, 6, f"Amount: Rupees {int(float(amount))} Only", 0, 1)

        # Clean up temp file
        if amman_path:
            try:
                import os
                os.unlink(amman_path)
            except:
                pass

        return bytes(pdf.output())

    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False

# ============================================================
# EXCEL ENGINE DETECTION
# ============================================================
EXCEL_ENGINE = None
try:
    import xlsxwriter
    EXCEL_ENGINE = 'xlsxwriter'
except ImportError:
    try:
        import openpyxl
        EXCEL_ENGINE = 'openpyxl'
    except ImportError:
        EXCEL_ENGINE = None

# ============================================================
# CONSTANTS
# ============================================================
NATCHATHIRAM_LIST = [
    "Ashwini", "Bharani", "Karthigai", "Rohini", "Mrigashirsha",
    "Thiruvadirai", "Punarvasu", "Pushya", "Ashlesha", "Magha",
    "Purva Phalguni", "Uttara Phalguni", "Hasta", "Chitra", "Swati",
    "Vishakha", "Anuradha", "Jyeshtha", "Mula", "Purva Ashadha",
    "Uttara Ashadha", "Shravana", "Dhanishta", "Shatabhisha",
    "Purva Bhadrapada", "Uttara Bhadrapada", "Revati"
]

RELATION_TYPES = [
    "Self", "Spouse", "Son", "Daughter", "Father", "Mother",
    "Brother", "Sister", "Grandfather", "Grandmother",
    "Father-in-law", "Mother-in-law", "Son-in-law",
    "Daughter-in-law", "Uncle", "Aunt", "Nephew", "Niece", "Other"
]

MIN_DATE = date(1900, 1, 1)
MAX_DATE = date(2050, 12, 31)

# ============================================================
# DEFAULT AMMAN SVG (Permanent fallback)
# ============================================================
DEFAULT_AMMAN_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 300 300" width="300" height="300">
<defs>
<radialGradient id="glow" cx="50%" cy="50%" r="50%">
<stop offset="0%" style="stop-color:#fff8f0;stop-opacity:1"/>
<stop offset="60%" style="stop-color:#ffe0b2;stop-opacity:1"/>
<stop offset="100%" style="stop-color:#ffcc80;stop-opacity:1"/>
</radialGradient>
<radialGradient id="inner" cx="50%" cy="45%" r="45%">
<stop offset="0%" style="stop-color:#fff3e0;stop-opacity:1"/>
<stop offset="100%" style="stop-color:#ffe0b2;stop-opacity:1"/>
</radialGradient>
</defs>
<circle cx="150" cy="150" r="148" fill="url(#glow)" stroke="#ff6b35" stroke-width="4"/>
<circle cx="150" cy="150" r="138" fill="url(#inner)" stroke="#f7c948" stroke-width="2"/>
<text x="150" y="55" text-anchor="middle" font-size="16" fill="#c62828" font-weight="bold">Om Amman</text>
<text x="150" y="95" text-anchor="middle" font-size="52">🙏</text>
<text x="150" y="130" text-anchor="middle" font-size="40">🪷</text>
<text x="150" y="165" text-anchor="middle" font-size="15" fill="#8B0000" font-weight="bold">Sree Bhadreshwari</text>
<text x="150" y="185" text-anchor="middle" font-size="15" fill="#8B0000" font-weight="bold">Amman</text>
<text x="150" y="210" text-anchor="middle" font-size="10" fill="#c62828">Amme Narayana</text>
<text x="150" y="235" text-anchor="middle" font-size="9" fill="#e65100">Devi Narayana</text>
</svg>"""

AMMAN_IMAGE_BASE64 = "data:image/svg+xml;base64," + base64.b64encode(
    DEFAULT_AMMAN_SVG.strip().encode()
).decode()


# ============================================================
# CUSTOM CSS
# ============================================================
def get_custom_css(amman_bg_url=None):
    bg_style = ""
    if amman_bg_url and not amman_bg_url.startswith('data:image/svg'):
        bg_style = f"""
        .login-bg-image {{
            position: fixed; top: 0; left: 0; width: 100%; height: 100%;
            background-image: url('{amman_bg_url}');
            background-size: cover; background-position: center;
            opacity: 0.08; z-index: -1; pointer-events: none;
        }}
        """

    return f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
        * {{ font-family: 'Poppins', sans-serif; }}

        {bg_style}

        .main-header {{
            background: linear-gradient(135deg, #ff6b35 0%, #f7c948 50%, #ff6b35 100%);
            padding: 15px 20px; border-radius: 15px; text-align: center;
            margin-bottom: 20px; box-shadow: 0 4px 15px rgba(255,107,53,0.3);
            position: relative; overflow: hidden;
        }}
        .main-header h1 {{ color: #8B0000; font-size: 1.6em; margin: 0; }}
        .main-header p {{ color: #5a1a00; font-size: 0.95em; margin: 5px 0 0 0; }}
        .header-amman-left, .header-amman-right {{
            position: absolute; top: 50%; transform: translateY(-50%);
            width: 50px; height: 50px; border-radius: 50%;
            border: 2px solid rgba(139,0,0,0.3);
            object-fit: cover;
        }}
        .header-amman-left {{ left: 15px; }}
        .header-amman-right {{ right: 15px; }}

        .login-container {{
            padding: 35px; border-radius: 20px;
            background: rgba(255,255,255,0.95);
            box-shadow: 0 10px 40px rgba(0,0,0,0.12);
            border: 2px solid rgba(255,107,53,0.15);
            position: relative; z-index: 1;
        }}

        .amman-circle {{ text-align: center; margin: 0 auto 20px auto; }}
        .amman-circle img {{
            width: 160px; height: 160px; border-radius: 50%;
            object-fit: cover; border: 5px solid #ff6b35;
            box-shadow: 0 0 25px rgba(255,107,53,0.4),
                        0 0 50px rgba(247,201,72,0.2),
                        0 0 75px rgba(255,107,53,0.1);
            animation: amman-glow 3s ease-in-out infinite alternate;
        }}
        @keyframes amman-glow {{
            0% {{ box-shadow: 0 0 25px rgba(255,107,53,0.4), 0 0 50px rgba(247,201,72,0.2); border-color: #ff6b35; }}
            50% {{ box-shadow: 0 0 35px rgba(255,107,53,0.6), 0 0 70px rgba(247,201,72,0.3), 0 0 100px rgba(255,107,53,0.15); border-color: #f7c948; }}
            100% {{ box-shadow: 0 0 25px rgba(255,107,53,0.4), 0 0 50px rgba(247,201,72,0.2); border-color: #ff6b35; }}
        }}

        .sidebar-amman {{ text-align: center; margin: 0 auto 10px auto; }}
        .sidebar-amman img {{
            width: 80px; height: 80px; border-radius: 50%;
            border: 3px solid #ff6b35;
            box-shadow: 0 0 15px rgba(255,107,53,0.3);
            object-fit: cover;
        }}

        .metric-card {{
            padding: 20px; border-radius: 12px; color: white;
            text-align: center; margin: 5px; box-shadow: 0 4px 10px rgba(0,0,0,0.15);
        }}
        .metric-card.income {{ background: linear-gradient(135deg, #11998e, #38ef7d); }}
        .metric-card.expense {{ background: linear-gradient(135deg, #eb3349, #f45c43); }}
        .metric-card.balance {{ background: linear-gradient(135deg, #4facfe, #00f2fe); }}
        .metric-card.info {{ background: linear-gradient(135deg, #667eea, #764ba2); }}
        .metric-card h3 {{ margin: 0; font-size: 0.85em; opacity: 0.9; }}
        .metric-card h2 {{ margin: 5px 0 0 0; font-size: 1.7em; }}

        .news-ticker-wrapper {{
            background: linear-gradient(90deg, #1a1a2e, #16213e, #0f3460);
            padding: 12px 20px; border-radius: 10px; overflow: hidden;
            white-space: nowrap; margin: 10px 0;
        }}
        .news-ticker-text {{
            display: inline-block; color: #f7c948; font-size: 1em;
            animation: scroll-left 35s linear infinite;
        }}
        @keyframes scroll-left {{
            0% {{ transform: translateX(100%); }}
            100% {{ transform: translateX(-200%); }}
        }}

        .pooja-card {{
            background: linear-gradient(135deg, #ffecd2, #fcb69f);
            padding: 12px 15px; border-radius: 10px; margin: 5px 0;
            border-left: 4px solid #ff6b35;
        }}
        .birthday-card {{
            background: linear-gradient(135deg, #a8edea, #fed6e3);
            padding: 10px 15px; border-radius: 10px; margin: 5px 0;
            border-left: 4px solid #e91e63;
        }}
        .success-box {{
            background: #d4edda; border: 1px solid #c3e6cb; padding: 15px;
            border-radius: 10px; color: #155724; margin: 10px 0;
        }}
        .wa-btn {{
            display: inline-block; background: #25D366; color: white !important;
            padding: 10px 25px; border-radius: 8px; text-decoration: none;
            font-weight: 600; font-size: 0.95em; margin: 5px;
            box-shadow: 0 3px 8px rgba(37,211,102,0.3);
        }}
        .wa-btn:hover {{ background: #128C7E; color: white !important; }}
        .upload-error {{
            background: #ffebee; border: 1px solid #ef9a9a; padding: 10px;
            border-radius: 8px; margin: 5px 0;
        }}

        div[data-testid="stSidebar"] {{
            background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
        }}
        div[data-testid="stSidebar"] .stButton > button {{
            width: 100%; text-align: left; background: transparent;
            color: #f0f0f0; border: 1px solid rgba(255,255,255,0.1);
            border-radius: 8px; margin: 2px 0; padding: 8px 15px;
        }}
        div[data-testid="stSidebar"] .stButton > button:hover {{
            background: rgba(255,107,53,0.3); border-color: #ff6b35;
        }}

        .temple-name-login {{
            color: #8B0000; font-size: 1.4em; font-weight: 700;
            text-align: center; margin: 10px 0; line-height: 1.3;
        }}
        .tamil-text {{
            color: #c0392b; font-size: 1.1em; font-weight: 600;
            text-align: center; margin: 5px 0 20px 0;
        }}

        .barcode-container {{
            background: white; padding: 15px; border-radius: 10px;
            border: 2px dashed #ccc; text-align: center; margin: 10px 0;
        }}
        .barcode-container img {{
            max-width: 100%; height: auto;
        }}
        .asset-card {{
            background: linear-gradient(135deg, #e8eaf6, #c5cae9);
            padding: 15px; border-radius: 12px; margin: 8px 0;
            border-left: 4px solid #3f51b5;
        }}
    </style>
    """


# ============================================================
# SESSION STATE
# ============================================================
defaults = {
    'logged_in': False, 'username': '', 'user_role': '',
    'current_page': 'Dashboard',
    'custom_amman_photo': None,
    'amman_photo_db': None,
}
for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val

# ============================================================
# AMMAN IMAGE MANAGEMENT
# ============================================================
def get_amman_image():
    """Returns the Amman image - from DB > session > default SVG"""
    # Priority 1: Database stored image
    if DB_CONNECTED:
        try:
            settings = db_select("temple_settings", filters={"key": "amman_image"})
            if settings and settings[0].get('value'):
                return settings[0]['value']
        except:
            pass

    # Priority 2: Session state
    if st.session_state.get('custom_amman_photo'):
        return st.session_state['custom_amman_photo']

    # Priority 3: Default SVG
    return AMMAN_IMAGE_BASE64


def save_amman_image_to_db(base64_img):
    """Save amman image to database for persistence"""
    if not DB_CONNECTED:
        return False
    try:
        existing = db_select("temple_settings", filters={"key": "amman_image"})
        if existing:
            db_update("temple_settings", {"value": base64_img}, "key", "amman_image")
        else:
            db_insert("temple_settings", {"key": "amman_image", "value": base64_img})
        st.session_state['custom_amman_photo'] = base64_img
        return True
    except:
        st.session_state['custom_amman_photo'] = base64_img
        return False


def get_amman_for_pdf():
    """Get amman image suitable for PDF (not SVG)"""
    img = get_amman_image()
    if img and not img.startswith('data:image/svg'):
        return img
    return None


# ============================================================
# BARCODE GENERATION FUNCTIONS
# ============================================================
def generate_barcode_image(data_str, barcode_type='code128'):
    """Generate barcode as base64 image"""
    if BARCODE_AVAILABLE:
        try:
            barcode_class = barcode.get_barcode_class(barcode_type)
            buffer = io.BytesIO()
            b = barcode_class(str(data_str), writer=ImageWriter())
            b.write(buffer, options={
                'module_width': 0.4,
                'module_height': 15,
                'font_size': 10,
                'text_distance': 5,
                'quiet_zone': 6.5
            })
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            return f"data:image/png;base64,{img_base64}", buffer.getvalue()
        except Exception:
            pass

    if QRCODE_AVAILABLE:
        try:
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(str(data_str))
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            return f"data:image/png;base64,{img_base64}", buffer.getvalue()
        except:
            pass

    # SVG fallback barcode
    return generate_svg_barcode(data_str), None


def generate_svg_barcode(data_str):
    """Generate a simple SVG barcode representation"""
    import hashlib
    hash_val = hashlib.md5(str(data_str).encode()).hexdigest()

    bars = []
    x = 10
    for i, char in enumerate(hash_val[:32]):
        val = int(char, 16)
        width = 2 if val > 7 else 1
        if val % 2 == 0:
            bars.append(f'<rect x="{x}" y="10" width="{width}" height="60" fill="black"/>')
        x += width + 1

    total_width = x + 10
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {total_width} 90" width="{total_width}" height="90">
    <rect width="{total_width}" height="90" fill="white"/>
    {''.join(bars)}
    <text x="{total_width/2}" y="82" text-anchor="middle" font-size="8" font-family="monospace">{data_str}</text>
    </svg>"""

    return "data:image/svg+xml;base64," + base64.b64encode(svg.encode()).decode()


def generate_asset_barcode_pdf(asset_tag, asset_name, barcode_img_bytes=None):
    """Generate a printable barcode label PDF"""
    if not PDF_AVAILABLE:
        return None

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'Sree Bhadreshwari Amman Temple', 0, 1, 'C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Asset Barcode Label', 0, 1, 'C')
    pdf.ln(10)

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 8, f'Asset Tag: {asset_tag}', 0, 1, 'C')
    pdf.cell(0, 8, f'Asset Name: {asset_name}', 0, 1, 'C')
    pdf.ln(5)

    if barcode_img_bytes:
        try:
            import tempfile
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            tmp.write(barcode_img_bytes)
            tmp.close()
            pdf.image(tmp.name, x=40, w=130)
            import os
            os.unlink(tmp.name)
        except:
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 8, f'Barcode: {asset_tag}', 0, 1, 'C')
    else:
        pdf.set_font('Courier', 'B', 16)
        pdf.cell(0, 10, f'|| {asset_tag} ||', 0, 1, 'C')

    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 5, f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}', 0, 1, 'C')

    return bytes(pdf.output())


# ============================================================
# DATABASE HELPERS
# ============================================================
def db_select(table, columns="*", filters=None, gte_filters=None, lte_filters=None):
    try:
        query = supabase.table(table).select(columns)
        if filters:
            for k, v in filters.items():
                query = query.eq(k, v)
        if gte_filters:
            for k, v in gte_filters.items():
                query = query.gte(k, str(v))
        if lte_filters:
            for k, v in lte_filters.items():
                query = query.lte(k, str(v))
        result = query.execute()
        return result.data if result.data else []
    except Exception:
        return []


def db_insert(table, data):
    try:
        result = supabase.table(table).insert(data).execute()
        return result.data if result.data else None
    except Exception as e:
        st.error(f"Insert Error ({table}): {e}")
        return None


def db_update(table, data, col, val):
    try:
        return supabase.table(table).update(data).eq(col, val).execute().data
    except Exception:
        return None


def db_delete(table, col, val):
    try:
        supabase.table(table).delete().eq(col, val).execute()
        return True
    except Exception:
        return False


def file_to_base64(f):
    if f:
        return f"data:{f.type};base64,{base64.b64encode(f.getvalue()).decode()}"
    return None


def get_income(s, e):
    return sum(
        float(b.get('amount', 0))
        for b in db_select("bills", "amount",
                           gte_filters={"bill_date": s},
                           lte_filters={"bill_date": e})
    )


def get_expense(s, e):
    return sum(
        float(x.get('amount', 0))
        for x in db_select("expenses", "amount",
                           gte_filters={"expense_date": s},
                           lte_filters={"expense_date": e})
    )


def get_period_dates(p):
    t = date.today()
    if p == "Daily":
        return t, t
    elif p == "Weekly":
        return t - timedelta(days=t.weekday()), t
    elif p == "Monthly":
        return t.replace(day=1), t
    elif p == "Yearly":
        return t.replace(month=1, day=1), t
    return t, t


def get_todays_birthdays():
    t = date.today()
    bdays = []
    for d in db_select("devotees", "name, dob"):
        if d.get('dob'):
            try:
                dob = datetime.strptime(str(d['dob']), '%Y-%m-%d').date()
                if dob.month == t.month and dob.day == t.day:
                    bdays.append(f"🎂 {d['name']} (Devotee)")
            except:
                pass
    for m in db_select("family_members", "name, dob"):
        if m.get('dob'):
            try:
                dob = datetime.strptime(str(m['dob']), '%Y-%m-%d').date()
                if dob.month == t.month and dob.day == t.day:
                    bdays.append(f"🎂 {m['name']} (Family)")
            except:
                pass
    return bdays


def gen_bill_no():
    return f"TMS-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"


def make_whatsapp_link(phone, message):
    phone_clean = ''.join(filter(str.isdigit, str(phone)))
    if len(phone_clean) == 10:
        phone_clean = "91" + phone_clean
    return f"https://wa.me/{phone_clean}?text={urllib.parse.quote(message)}"


def parse_date_safe(val):
    if val is None or str(val).strip() == '' or str(val).lower() in ('nan', 'nat', 'none'):
        return None
    val_str = str(val).strip()
    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
                '%Y/%m/%d', '%d-%m-%y', '%d/%m/%y']:
        try:
            return datetime.strptime(val_str, fmt).date()
        except:
            pass
    try:
        ts = pd.Timestamp(val)
        if not pd.isna(ts):
            return ts.date()
    except:
        pass
    return None


def safe_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan', 'none', 'nat') else s


# ============================================================
# BANNER WITH AMMAN IMAGE
# ============================================================
def render_page_header(title, subtitle=""):
    """Render page header with Amman image on both sides"""
    amman_img = get_amman_image()
    st.markdown(f"""
    <div class="main-header">
        <img src="{amman_img}" class="header-amman-left" alt="">
        <h1>{title}</h1>
        <p>{subtitle}</p>
        <img src="{amman_img}" class="header-amman-right" alt="">
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# EXCEL TEMPLATE
# ============================================================
def generate_bulk_template():
    columns = ['Sl_No', 'Type', 'Family_Head_Name', 'Member_Name', 'Address',
               'Mobile_No', 'WhatsApp_No', 'Relation_Type', 'Date_of_Birth',
               'Natchathiram', 'Wedding_Day', 'Yearly_Pooja', 'Yearly_Pooja_Dates']
    sample = [
        ['1', 'HEAD', 'Raman K', '', '12 Main St', '9876543210',
         '9876543210', 'Self', '15-05-1980', 'Ashwini', '10-06-2005',
         'Archana;Abhishekam', '15-01-2025;20-06-2025'],
        ['2', 'HEAD', 'Suresh M', '', '45 Temple Rd', '9876543211',
         '9876543211', 'Self', '20-08-1975', 'Rohini', '15-01-2000',
         'Homam', '10-03-2025'],
        ['1.1', 'MEMBER', 'Raman K', 'Lakshmi R', '', '', '',
         'Spouse', '20-07-1985', 'Bharani', '10-06-2005', '', ''],
        ['1.2', 'MEMBER', 'Raman K', 'Karthik R', '', '', '',
         'Son', '10-03-2008', 'Rohini', '', '', ''],
        ['2.1', 'MEMBER', 'Suresh M', 'Priya S', '', '', '',
         'Spouse', '25-12-1980', 'Magha', '15-01-2000', '', ''],
    ]
    df = pd.DataFrame(sample, columns=columns)

    if EXCEL_ENGINE:
        try:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine=EXCEL_ENGINE) as writer:
                df.to_excel(writer, index=False, sheet_name='Devotees')
                instr = pd.DataFrame({'Instructions': [
                    'BULK UPLOAD TEMPLATE', '',
                    'Type: HEAD or MEMBER',
                    'Family_Head_Name: Links members to head',
                    'Member_Name: Members own name',
                    'Date format: DD-MM-YYYY',
                    'Multiple poojas: separate with ;',
                    '', 'Stars: ' + ', '.join(NATCHATHIRAM_LIST),
                    '', 'Relations: ' + ', '.join(RELATION_TYPES),
                ]})
                instr.to_excel(writer, index=False, sheet_name='Instructions')
            return (output.getvalue(), 'devotee_template.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception:
            pass

    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode('utf-8'), 'devotee_template.csv', 'text/csv'


def process_bulk_upload(df):
    results = {'success': 0, 'errors': [], 'members_added': 0, 'poojas_added': 0}
    head_id_map = {}
    df.columns = [c.strip().replace(' ', '_') for c in df.columns]

    for col in ['Type', 'Family_Head_Name']:
        if col not in df.columns:
            results['errors'].append(f"Missing column: {col}")
            return results

    heads = df[df['Type'].astype(str).str.upper().str.strip() == 'HEAD']
    for idx, row in heads.iterrows():
        try:
            name = safe_str(row.get('Family_Head_Name'))
            if not name:
                results['errors'].append(f"Row {idx + 2}: No name")
                continue
            dob = parse_date_safe(row.get('Date_of_Birth'))
            wed = parse_date_safe(row.get('Wedding_Day'))
            r = db_insert("devotees", {
                "name": name,
                "dob": str(dob) if dob else None,
                "relation_type": safe_str(row.get('Relation_Type')) or 'Self',
                "mobile_no": safe_str(row.get('Mobile_No')),
                "whatsapp_no": safe_str(row.get('WhatsApp_No')),
                "wedding_day": str(wed) if wed else None,
                "natchathiram": safe_str(row.get('Natchathiram')) or None,
                "address": safe_str(row.get('Address')),
            })
            if r:
                hid = r[0]['id']
                head_id_map[name.lower().strip()] = hid
                results['success'] += 1
                ps = safe_str(row.get('Yearly_Pooja'))
                ds = safe_str(row.get('Yearly_Pooja_Dates'))
                if ps:
                    for i, pn in enumerate(
                            [p.strip() for p in ps.split(';') if p.strip()]):
                        pd_list = ([d.strip() for d in ds.split(';')
                                    if d.strip()] if ds else [])
                        pd_val = (parse_date_safe(pd_list[i])
                                  if i < len(pd_list) else None)
                        db_insert("devotee_yearly_pooja", {
                            "devotee_id": hid, "pooja_type": pn,
                            "pooja_date": str(pd_val) if pd_val else None,
                            "description": "Bulk"
                        })
                        results['poojas_added'] += 1
        except Exception as e:
            results['errors'].append(f"Row {idx + 2}: {e}")

    members = df[df['Type'].astype(str).str.upper().str.strip() == 'MEMBER']
    for idx, row in members.iterrows():
        try:
            href = safe_str(row.get('Family_Head_Name')).lower().strip()
            mname = (safe_str(row.get('Member_Name'))
                     or safe_str(row.get('Address'))
                     or f"Member of {href}")
            hid = head_id_map.get(href)
            if not hid:
                for d in db_select("devotees", "id, name"):
                    if d['name'].lower().strip() == href:
                        hid = d['id']
                        break
            if not hid:
                results['errors'].append(f"Row {idx + 2}: Head not found")
                continue
            dob = parse_date_safe(row.get('Date_of_Birth'))
            wed = parse_date_safe(row.get('Wedding_Day'))
            if db_insert("family_members", {
                "devotee_id": hid, "name": mname,
                "dob": str(dob) if dob else None,
                "relation_type": safe_str(row.get('Relation_Type')),
                "wedding_day": str(wed) if wed else None,
                "natchathiram": safe_str(row.get('Natchathiram')) or None
            }):
                results['members_added'] += 1
        except Exception as e:
            results['errors'].append(f"Row {idx + 2}: {e}")

    return results


# ============================================================
# PAGE: LOGIN
# ============================================================
def page_login():
    amman_img = get_amman_image()

    # Apply CSS with background
    st.markdown(get_custom_css(amman_img), unsafe_allow_html=True)

    # Background overlay for login
    if amman_img and not amman_img.startswith('data:image/svg'):
        st.markdown(f'<div class="login-bg-image"></div>', unsafe_allow_html=True)

    st.markdown("""
    <style>
        .stApp {{
            background: linear-gradient(
                135deg, #fff5ee 0%, #ffe4c4 25%,
                #ffdab9 50%, #ffe4c4 75%, #fff5ee 100%
            );
        }}
    </style>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown('<div class="login-container">', unsafe_allow_html=True)

        # AMMAN PHOTO IN ROUND
        st.markdown(f"""
        <div class="amman-circle">
            <img src="{amman_img}" alt="Sree Bhadreshwari Amman">
        </div>
        """, unsafe_allow_html=True)

        # Temple Name
        st.markdown("""
        <div class="temple-name-login">
            🛕 Sree Bhadreshwari Amman Temple<br>Management System
        </div>
        <div class="tamil-text">
            🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏
        </div>
        """, unsafe_allow_html=True)

        with st.form("login"):
            u = st.text_input("👤 Username")
            p = st.text_input("🔑 Password", type="password")
            if st.form_submit_button("🚀 Login", use_container_width=True):
                if not u or not p:
                    st.warning("⚠️ Enter both fields!")
                elif not DB_CONNECTED:
                    st.error("❌ DB not connected!")
                else:
                    users = db_select("users", filters={"username": u})
                    if users and users[0].get('password_hash') == p:
                        st.session_state.logged_in = True
                        st.session_state.username = u
                        st.session_state.user_role = users[0].get('role', 'user')
                        st.success("✅ Success!")
                        time.sleep(0.5)
                        st.rerun()
                    else:
                        st.error("❌ Invalid credentials!")

        st.markdown("""
        <div style="text-align:center;color:#999;font-size:0.8em;margin-top:15px;">
            Default Login: admin / admin123
        </div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
# PAGE: DASHBOARD
# ============================================================
def page_dashboard():
    render_page_header(
        "🛕 Sree Bhadreshwari Amman Temple",
        "🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏"
    )

    tparts = get_todays_birthdays()
    for n in db_select("news_ticker", filters={"is_active": True}):
        tparts.append(f"📢 {n['message']}")
    if not tparts:
        tparts.append("🛕 Welcome to Sree Bhadreshwari Amman Temple! 🙏")
    st.markdown(
        f'<div class="news-ticker-wrapper"><div class="news-ticker-text">'
        f'{" &nbsp;⭐&nbsp; ".join(tparts)}</div></div>',
        unsafe_allow_html=True
    )

    period = st.selectbox("📅 Period", ["Daily", "Weekly", "Monthly", "Yearly"])
    s, e = get_period_dates(period)
    inc, exp = get_income(s, e), get_expense(s, e)
    bal = inc - exp
    td = len(db_select("devotees", "id"))

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(
            f'<div class="metric-card income"><h3>💰 {period} Income</h3>'
            f'<h2>₹ {inc:,.2f}</h2></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(
            f'<div class="metric-card expense"><h3>💸 {period} Expenses</h3>'
            f'<h2>₹ {exp:,.2f}</h2></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(
            f'<div class="metric-card balance"><h3>💎 Balance</h3>'
            f'<h2>₹ {bal:,.2f}</h2></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(
            f'<div class="metric-card info"><h3>👥 Devotees</h3>'
            f'<h2>{td}</h2></div>', unsafe_allow_html=True)

    st.markdown("---")
    cl, cr = st.columns(2)
    with cl:
        st.markdown("### 🎂 Birthdays Today")
        bdays = get_todays_birthdays()
        for b in bdays:
            st.markdown(
                f'<div class="birthday-card">🎉 {b} 🎈</div>',
                unsafe_allow_html=True)
        if not bdays:
            st.info("No birthdays today")
    with cr:
        st.markdown("### 🙏 Today's Pooja")
        for p in db_select("daily_pooja", filters={"pooja_date": str(date.today())}):
            ic = "✅" if p.get('status') == 'completed' else "⏳"
            st.markdown(
                f'<div class="pooja-card">{ic} <b>{p["pooja_name"]}</b>'
                f' — {p.get("pooja_time", "")}</div>',
                unsafe_allow_html=True)
            if p.get('status') != 'completed':
                if st.button("Complete", key=f"c_{p['id']}"):
                    db_update("daily_pooja",
                              {"status": "completed"}, "id", p['id'])
                    st.rerun()
        with st.expander("➕ Add Pooja"):
            with st.form("adp"):
                dn = st.text_input("Name")
                dt_t = st.text_input("Time")
                dd = st.date_input("Date")
                if st.form_submit_button("Add"):
                    if dn:
                        db_insert("daily_pooja", {
                            "pooja_name": dn, "pooja_time": dt_t,
                            "pooja_date": str(dd), "status": "pending"
                        })
                        st.rerun()

    st.markdown("---")
    st.bar_chart(pd.DataFrame({
        "Category": ["Income", "Expenses", "Balance"],
        "₹": [inc, exp, bal]
    }).set_index("Category"))


# ============================================================
# PAGE: DEVOTEE ENROLLMENT
# ============================================================
def page_devotee_enrollment():
    render_page_header("👥 Devotee Enrollment",
                       "Register, Bulk Upload & Manage")

    tab1, tab2, tab3, tab4 = st.tabs(
        ["➕ New", "📤 Bulk Upload", "🔍 Search", "👨‍👩‍👧‍👦 Family"])

    with tab1:
        with st.form("enroll", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nm = st.text_input("👤 Name *")
                db_v = st.date_input("📅 DOB", value=date(1990, 1, 1),
                                     min_value=MIN_DATE, max_value=MAX_DATE)
                rl = st.selectbox("👪 Relation", RELATION_TYPES)
                mb = st.text_input("📱 Mobile")
                wa = st.text_input("📲 WhatsApp")
            with c2:
                wd = st.date_input("💒 Wedding", value=None,
                                   min_value=MIN_DATE, max_value=MAX_DATE)
                nt = st.selectbox("⭐ Star", ["--"] + NATCHATHIRAM_LIST)
                ad = st.text_area("🏠 Address", height=80)
                ph = st.file_uploader("📷 Photo",
                                      type=['jpg', 'jpeg', 'png'])
            st.markdown("#### 🙏 Yearly Pooja")
            yc1, yc2, yc3 = st.columns(3)
            ptl = [p['name'] for p in db_select("pooja_types", "name")]
            with yc1:
                ypt = st.selectbox("Type", ["--"] + ptl, key="y1t")
            with yc2:
                ypd = st.date_input("Date", key="y1d",
                                    min_value=MIN_DATE, max_value=MAX_DATE)
            with yc3:
                ypdesc = st.text_input("Desc", key="y1dc")
            if st.form_submit_button("✅ Register", use_container_width=True):
                if nm.strip():
                    r = db_insert("devotees", {
                        "name": nm.strip(), "dob": str(db_v),
                        "relation_type": rl, "mobile_no": mb,
                        "whatsapp_no": wa,
                        "wedding_day": str(wd) if wd else None,
                        "natchathiram": nt if nt != "--" else None,
                        "address": ad, "photo_url": file_to_base64(ph)
                    })
                    if r and ypt != "--":
                        db_insert("devotee_yearly_pooja", {
                            "devotee_id": r[0]['id'], "pooja_type": ypt,
                            "pooja_date": str(ypd), "description": ypdesc
                        })
                    if r:
                        st.success(f"✅ '{nm}' enrolled!")
                        st.rerun()

    with tab2:
        st.markdown("### 📤 Bulk Upload Devotees")
        tb, tn, tm = generate_bulk_template()
        st.download_button(
            f"📥 Download Template ({tn.split('.')[-1].upper()})",
            data=tb, file_name=tn, mime=tm, use_container_width=True)
        st.markdown("""
        **Format:** Type=HEAD/MEMBER | Family_Head_Name links members  
        **Dates:** DD-MM-YYYY | **Multiple poojas:** separate with `;`
        """)
        uf = st.file_uploader("📁 Upload File",
                              type=['xlsx', 'xls', 'csv'], key="bulk")
        if uf:
            try:
                df = (pd.read_csv(uf) if uf.name.endswith('.csv')
                      else pd.read_excel(uf, sheet_name=0))
                st.dataframe(df.head(15), use_container_width=True,
                             hide_index=True)
                if st.button("🚀 Process & Upload",
                             use_container_width=True, type="primary"):
                    with st.spinner("Processing..."):
                        res = process_bulk_upload(df)
                    rc1, rc2, rc3 = st.columns(3)
                    with rc1:
                        st.markdown(
                            f'<div class="metric-card income"><h3>Heads</h3>'
                            f'<h2>{res["success"]}</h2></div>',
                            unsafe_allow_html=True)
                    with rc2:
                        st.markdown(
                            f'<div class="metric-card balance"><h3>Members</h3>'
                            f'<h2>{res["members_added"]}</h2></div>',
                            unsafe_allow_html=True)
                    with rc3:
                        st.markdown(
                            f'<div class="metric-card info"><h3>Poojas</h3>'
                            f'<h2>{res["poojas_added"]}</h2></div>',
                            unsafe_allow_html=True)
                    if res['errors']:
                        with st.expander(f"⚠️ {len(res['errors'])} Errors"):
                            for err in res['errors']:
                                st.markdown(
                                    f'<div class="upload-error">❌ {err}</div>',
                                    unsafe_allow_html=True)
                    if res['success'] > 0:
                        st.balloons()
            except Exception as e:
                st.error(f"Error: {e}")

    with tab3:
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            sn = st.text_input("Name", key="sn")
        with sc2:
            sm = st.text_input("Mobile", key="sm")
        with sc3:
            sa = st.text_input("Address", key="sa")
        devs = db_select("devotees")
        if sn:
            devs = [d for d in devs if sn.lower() in d.get('name', '').lower()]
        if sm:
            devs = [d for d in devs if sm in d.get('mobile_no', '')]
        if sa:
            devs = [d for d in devs
                    if sa.lower() in d.get('address', '').lower()]
        st.markdown(f"**Found: {len(devs)}**")
        for dev in devs:
            with st.expander(
                    f"👤 {dev['name']} | 📱 {dev.get('mobile_no', 'N/A')}"):
                dc1, dc2 = st.columns([3, 1])
                with dc1:
                    for l, k in [
                        ("Name", "name"), ("DOB", "dob"),
                        ("Mobile", "mobile_no"), ("WhatsApp", "whatsapp_no"),
                        ("Relation", "relation_type"),
                        ("Wedding", "wedding_day"),
                        ("Star", "natchathiram"), ("Address", "address")
                    ]:
                        st.write(f"**{l}:** {dev.get(k, 'N/A')}")
                with dc2:
                    if (dev.get('photo_url')
                            and dev['photo_url'].startswith('data:')):
                        st.markdown(
                            f'<img src="{dev["photo_url"]}" width="120" '
                            f'style="border-radius:10px">',
                            unsafe_allow_html=True)
                st.markdown("**Yearly Poojas:**")
                for yp in db_select("devotee_yearly_pooja",
                                    filters={"devotee_id": dev['id']}):
                    yc1, yc2 = st.columns([5, 1])
                    with yc1:
                        st.write(
                            f"• {yp['pooja_type']} — {yp.get('pooja_date', '')}")
                    with yc2:
                        if st.button("❌", key=f"dyp_{yp['id']}"):
                            db_delete("devotee_yearly_pooja", "id", yp['id'])
                            st.rerun()
                with st.form(f"ayp_{dev['id']}"):
                    ac1, ac2, ac3 = st.columns(3)
                    ptn = [p['name']
                           for p in db_select("pooja_types", "name")]
                    with ac1:
                        nypt = st.selectbox(
                            "Type", ["--"] + ptn, key=f"nt_{dev['id']}")
                    with ac2:
                        nypd = st.date_input(
                            "Date", key=f"nd_{dev['id']}",
                            min_value=MIN_DATE, max_value=MAX_DATE)
                    with ac3:
                        ndc = st.text_input("Desc", key=f"ndc_{dev['id']}")
                    if st.form_submit_button("Add"):
                        if nypt != "--":
                            db_insert("devotee_yearly_pooja", {
                                "devotee_id": dev['id'],
                                "pooja_type": nypt,
                                "pooja_date": str(nypd),
                                "description": ndc
                            })
                            st.rerun()
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("✏️", key=f"e_{dev['id']}"):
                        st.session_state[f"ed_{dev['id']}"] = \
                            not st.session_state.get(f"ed_{dev['id']}", False)
                        st.rerun()
                with bc2:
                    if st.button("🗑️", key=f"d_{dev['id']}"):
                        db_delete("devotee_yearly_pooja",
                                  "devotee_id", dev['id'])
                        db_delete("family_members", "devotee_id", dev['id'])
                        db_delete("devotees", "id", dev['id'])
                        st.rerun()
                if st.session_state.get(f"ed_{dev['id']}", False):
                    with st.form(f"ef_{dev['id']}"):
                        ec1, ec2 = st.columns(2)
                        with ec1:
                            en = st.text_input(
                                "Name", value=dev.get('name', ''),
                                key=f"en_{dev['id']}")
                            edv = date(1990, 1, 1)
                            try:
                                edv = datetime.strptime(
                                    str(dev.get('dob', '')),
                                    "%Y-%m-%d").date()
                            except:
                                pass
                            ed = st.date_input(
                                "DOB", value=edv, key=f"ed2_{dev['id']}",
                                min_value=MIN_DATE, max_value=MAX_DATE)
                            em = st.text_input(
                                "Mobile",
                                value=dev.get('mobile_no', ''),
                                key=f"em_{dev['id']}")
                        with ec2:
                            er = st.selectbox(
                                "Relation", RELATION_TYPES,
                                index=(RELATION_TYPES.index(
                                    dev['relation_type'])
                                    if dev.get('relation_type')
                                    in RELATION_TYPES else 0),
                                key=f"er_{dev['id']}")
                            so = ["--"] + NATCHATHIRAM_LIST
                            cs = dev.get('natchathiram', '--')
                            es = st.selectbox(
                                "Star", so,
                                index=so.index(cs) if cs in so else 0,
                                key=f"es_{dev['id']}")
                            ea = st.text_area(
                                "Address", value=dev.get('address', ''),
                                key=f"ea_{dev['id']}")
                        if st.form_submit_button("💾"):
                            db_update("devotees", {
                                "name": en, "dob": str(ed),
                                "mobile_no": em, "relation_type": er,
                                "natchathiram": es if es != "--" else None,
                                "address": ea
                            }, "id", dev['id'])
                            st.session_state[f"ed_{dev['id']}"] = False
                            st.rerun()

    with tab4:
        ds = db_select("devotees", "id,name,mobile_no")
        if not ds:
            st.info("No devotees")
            return
        do = {f"{d['name']} ({d.get('mobile_no', '')})": d['id'] for d in ds}
        sh = st.selectbox("Head", list(do.keys()))
        hi = do[sh]
        for fm in db_select("family_members", filters={"devotee_id": hi}):
            fc1, fc2 = st.columns([5, 1])
            with fc1:
                st.write(
                    f"👤 **{fm['name']}** | {fm.get('relation_type', '')}"
                    f" | {fm.get('dob', '')}")
            with fc2:
                if st.button("🗑️", key=f"dfm_{fm['id']}"):
                    db_delete("family_members", "id", fm['id'])
                    st.rerun()
        with st.form("afm", clear_on_submit=True):
            fc1, fc2 = st.columns(2)
            with fc1:
                fn = st.text_input("Name *")
                fd = st.date_input("DOB", value=date(1995, 1, 1),
                                   min_value=MIN_DATE, max_value=MAX_DATE)
                fr = st.selectbox("Relation", RELATION_TYPES)
            with fc2:
                fw = st.date_input("Wedding", value=None,
                                   min_value=MIN_DATE, max_value=MAX_DATE,
                                   key="fmw")
                fs = st.selectbox("Star", ["--"] + NATCHATHIRAM_LIST,
                                  key="fms")
            if st.form_submit_button("➕", use_container_width=True):
                if fn.strip():
                    db_insert("family_members", {
                        "devotee_id": hi, "name": fn.strip(),
                        "dob": str(fd), "relation_type": fr,
                        "wedding_day": str(fw) if fw else None,
                        "natchathiram": fs if fs != "--" else None
                    })
                    st.rerun()


# ============================================================
# PAGE: BILLING (Enhanced with PDF + WhatsApp)
# ============================================================
def page_billing():
    render_page_header("🧾 Billing",
                       "PDF Download & WhatsApp Integration")

    tab1, tab2 = st.tabs(["➕ New Bill", "📋 History"])

    with tab1:
        dt = st.radio("Type", ["Enrolled", "Guest"], horizontal=True)
        bc1, bc2 = st.columns(2)
        with bc1:
            mbl = st.text_input("📝 Manual Bill No")
            bb = st.text_input("📖 Book No")
            ptd = db_select("pooja_types")
            pto = ({f"{p['name']} — ₹{p.get('amount', 0)}": p
                    for p in ptd} if ptd else {})
            sp = st.selectbox("🙏 Pooja",
                              list(pto.keys()) if pto else ["None"])
            da = (float(pto[sp].get('amount', 0))
                  if sp in pto else 0.0)
            am = st.number_input("💰 Amount", value=da,
                                 min_value=0.0, step=10.0)
            bd = st.date_input("📅 Date", value=date.today())
        with bc2:
            did = None
            gn = ga = gm = gw = ""
            if dt == "Enrolled":
                sby = st.selectbox(
                    "Search By",
                    ["Name", "Mobile", "WhatsApp", "Address"])
                sv = st.text_input(f"Enter {sby}")
                al = db_select("devotees")
                if sv:
                    fm = {"Name": "name", "Mobile": "mobile_no",
                          "WhatsApp": "whatsapp_no", "Address": "address"}
                    al = [d for d in al
                          if sv.lower() in str(
                              d.get(fm[sby], '')).lower()]
                if al:
                    dm = {f"{d['name']} — {d.get('mobile_no', 'N/A')}": d
                          for d in al}
                    ch = st.selectbox("Select Devotee", list(dm.keys()))
                    if ch:
                        sd = dm[ch]
                        did = sd['id']
                        st.markdown(
                            f'<div class="success-box">'
                            f'👤 <b>{sd["name"]}</b><br>'
                            f'📱 {sd.get("mobile_no", "N/A")} '
                            f'📲 {sd.get("whatsapp_no", "N/A")}<br>'
                            f'🏠 {sd.get("address", "N/A")}</div>',
                            unsafe_allow_html=True)
            else:
                gn = st.text_input("Name *")
                ga = st.text_area("Address *", height=60)
                gm = st.text_input("📱 Mobile")
                gw = st.text_input("📲 WhatsApp")

        if st.button("🧾 Generate Bill", use_container_width=True,
                     type="primary"):
            ok = True
            if dt == "Enrolled" and not did:
                st.error("Select a devotee!")
                ok = False
            if dt == "Guest" and not gn.strip():
                st.error("Enter name!")
                ok = False
            if am <= 0:
                st.error("Enter amount!")
                ok = False
            if ok:
                bn = gen_bill_no()
                pn = sp.split(" — ")[0] if " — " in sp else sp
                res = db_insert("bills", {
                    "bill_no": bn, "manual_bill_no": mbl,
                    "bill_book_no": bb,
                    "devotee_type": ("enrolled" if dt == "Enrolled"
                                     else "guest"),
                    "devotee_id": did,
                    "guest_name": gn if dt == "Guest" else None,
                    "guest_address": ga if dt == "Guest" else None,
                    "guest_mobile": gm if dt == "Guest" else None,
                    "guest_whatsapp": gw if dt == "Guest" else None,
                    "pooja_type": pn, "amount": am,
                    "bill_date": str(bd)
                })
                if res:
                    if dt == "Enrolled" and did:
                        di = db_select("devotees", filters={"id": did})
                        bn_ = di[0]['name'] if di else "N/A"
                        ba = di[0].get('address', '') if di else ""
                        bm = di[0].get('mobile_no', '') if di else ""
                        bwn = di[0].get('whatsapp_no', '') if di else ""
                    else:
                        bn_, ba, bm, bwn = gn, ga, gm, gw

                    st.success(f"✅ Bill Generated: {bn}")

                    # Bill display
                    amman_img = get_amman_image()
                    st.markdown(f"""
                    <div style="background:#fffdf7;padding:25px;border:2px solid #ff6b35;
                        border-radius:15px;max-width:600px;margin:20px auto;">
                        <div style="text-align:center;border-bottom:2px solid #ff6b35;
                            padding-bottom:12px;position:relative;">
                            <img src="{amman_img}" style="width:45px;height:45px;
                                border-radius:50%;position:absolute;left:10px;top:0;
                                border:2px solid #ff6b35;">
                            <h2 style="color:#8B0000;margin:0;">
                                🛕 Sree Bhadreshwari Amman Temple</h2>
                            <p style="margin:3px 0;">
                                🙏 அம்மே நாராயணா 🙏</p>
                            <img src="{amman_img}" style="width:45px;height:45px;
                                border-radius:50%;position:absolute;right:10px;top:0;
                                border:2px solid #ff6b35;">
                        </div>
                        <table style="width:100%;margin:15px 0;">
                            <tr><td><b>Bill:</b></td><td>{bn}</td></tr>
                            <tr><td><b>Manual:</b></td><td>{mbl}</td></tr>
                            <tr><td><b>Book:</b></td><td>{bb}</td></tr>
                            <tr><td><b>Date:</b></td><td>{bd}</td></tr>
                            <tr><td colspan="2">
                                <hr style="border:1px dashed #ccc"></td></tr>
                            <tr><td><b>Name:</b></td><td>{bn_}</td></tr>
                            <tr><td><b>Address:</b></td><td>{ba}</td></tr>
                            <tr><td><b>Mobile:</b></td><td>{bm}</td></tr>
                            <tr><td colspan="2">
                                <hr style="border:1px dashed #ccc"></td></tr>
                            <tr><td><b>Pooja:</b></td><td>{pn}</td></tr>
                            <tr><td><b>Amount:</b></td>
                                <td style="font-size:1.4em;color:#11998e">
                                <b>₹ {am:,.2f}</b></td></tr>
                        </table>
                        <div style="text-align:center;border-top:2px solid #ff6b35;
                            padding-top:10px;">
                            <p style="color:#666;margin:0;">
                                🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏</p>
                        </div>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("---")

                    # Download & WhatsApp buttons
                    dl1, dl2, dl3 = st.columns(3)

                    with dl1:
                        if PDF_AVAILABLE:
                            try:
                                amman_pdf = get_amman_for_pdf()
                                pdf = generate_bill_pdf(
                                    bn, mbl, bb, bd, bn_, ba, bm, pn, am,
                                    amman_base64=amman_pdf)
                                st.download_button(
                                    "📥 Download PDF", data=pdf,
                                    file_name=f"Bill_{bn}.pdf",
                                    mime="application/pdf",
                                    use_container_width=True)
                            except Exception as ex:
                                st.warning(f"PDF error: {ex}")
                        else:
                            st.info("📄 PDF not available")

                    with dl2:
                        wn = bwn or bm
                        if wn:
                            msg = (
                                f"🛕 *Sree Bhadreshwari Amman Temple*\n"
                                f"🙏 அம்மே நாராயணா\n\n"
                                f"📋 *BILL RECEIPT*\n"
                                f"━━━━━━━━━━━━━━━\n"
                                f"Bill No: {bn}\n"
                                f"Date: {bd}\n"
                                f"Name: {bn_}\n"
                                f"Pooja: {pn}\n"
                                f"*Amount: ₹ {am:,.2f}*\n"
                                f"━━━━━━━━━━━━━━━\n\n"
                                f"🙏 Thank you for your contribution!\n"
                                f"May Goddess bless you!\n"
                                f"🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏"
                            )
                            wa_link = make_whatsapp_link(wn, msg)
                            st.markdown(
                                f'<a href="{wa_link}" target="_blank" '
                                f'class="wa-btn">📲 Send WhatsApp</a>',
                                unsafe_allow_html=True)
                        else:
                            st.info("No WhatsApp number")

                    with dl3:
                        # Copy bill text
                        bill_text = (
                            f"Bill: {bn}\nDate: {bd}\n"
                            f"Name: {bn_}\nPooja: {pn}\n"
                            f"Amount: Rs.{am:,.2f}")
                        st.code(bill_text, language=None)

    with tab2:
        st.markdown("### 📋 Bill History")

        # Filter options
        hf1, hf2, hf3 = st.columns(3)
        with hf1:
            h_from = st.date_input("From", value=date.today() - timedelta(30),
                                   key="hf_from")
        with hf2:
            h_to = st.date_input("To", value=date.today(), key="hf_to")
        with hf3:
            h_search = st.text_input("🔍 Search", key="hf_search")

        all_bills = sorted(
            db_select("bills", gte_filters={"bill_date": h_from},
                      lte_filters={"bill_date": h_to}),
            key=lambda x: x.get('created_at', ''), reverse=True)

        for b in all_bills:
            bname = b.get('guest_name', '')
            bwn = (b.get('guest_whatsapp', '')
                   or b.get('guest_mobile', ''))
            baddr = b.get('guest_address', '')
            bmob = b.get('guest_mobile', '')

            if (b.get('devotee_type') == 'enrolled'
                    and b.get('devotee_id')):
                dd = db_select(
                    "devotees",
                    "name,mobile_no,whatsapp_no,address",
                    filters={"id": b['devotee_id']})
                if dd:
                    bname = dd[0]['name']
                    bmob = dd[0].get('mobile_no', '')
                    bwn = (dd[0].get('whatsapp_no', '')
                           or dd[0].get('mobile_no', ''))
                    baddr = dd[0].get('address', '')

            # Apply search filter
            if h_search:
                search_in = f"{bname} {b.get('bill_no', '')} {b.get('pooja_type', '')}".lower()
                if h_search.lower() not in search_in:
                    continue

            with st.expander(
                    f"🧾 {b.get('bill_no', '')} | {bname} "
                    f"| ₹{b.get('amount', 0)} | {b.get('bill_date', '')}"):
                hc1, hc2, hc3, hc4 = st.columns(4)

                with hc1:
                    if PDF_AVAILABLE:
                        try:
                            amman_pdf = get_amman_for_pdf()
                            pdf_data = generate_bill_pdf(
                                b.get('bill_no', ''),
                                b.get('manual_bill_no', ''),
                                b.get('bill_book_no', ''),
                                b.get('bill_date', ''),
                                bname, baddr, bmob,
                                b.get('pooja_type', ''),
                                b.get('amount', 0),
                                amman_base64=amman_pdf)
                            st.download_button(
                                "📥 PDF", data=pdf_data,
                                file_name=f"Bill_{b.get('bill_no', '')}.pdf",
                                mime="application/pdf",
                                key=f"p_{b['id']}")
                        except:
                            st.info("PDF N/A")

                with hc2:
                    if bwn:
                        msg = (
                            f"🛕 *Sree Bhadreshwari Amman Temple*\n"
                            f"Bill: {b.get('bill_no', '')}\n"
                            f"Date: {b.get('bill_date', '')}\n"
                            f"Name: {bname}\n"
                            f"Pooja: {b.get('pooja_type', '')}\n"
                            f"*Amount: ₹{b.get('amount', 0):,.2f}*\n"
                            f"🙏 Thank you!")
                        wa_link = make_whatsapp_link(bwn, msg)
                        st.markdown(
                            f'<a href="{wa_link}" target="_blank" '
                            f'class="wa-btn" style="font-size:0.8em;'
                            f'padding:5px 10px">📲 WhatsApp</a>',
                            unsafe_allow_html=True)

                with hc3:
                    st.write(f"**Pooja:** {b.get('pooja_type', '')}")
                    st.write(f"**Amount:** ₹{b.get('amount', 0):,.2f}")

                with hc4:
                    if st.session_state.user_role == 'admin':
                        if st.button("🗑️", key=f"db_{b['id']}"):
                            db_delete("bills", "id", b['id'])
                            st.rerun()


# ============================================================
# PAGE: EXPENSES
# ============================================================
def page_expenses():
    render_page_header("💸 Expenses", "Track temple expenses")

    t1, t2 = st.tabs(["➕ Add", "📋 History"])
    with t1:
        with st.form("ef", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                etn = ([e['name'] for e in db_select("expense_types", "name")]
                       or ["Misc"])
                et = st.selectbox("Type", etn)
                ea = st.number_input("Amount", min_value=0.0, step=10.0)
            with c2:
                ed = st.date_input("Date")
                edesc = st.text_area("Description", height=80)
            if st.form_submit_button("💾", use_container_width=True):
                if ea > 0:
                    db_insert("expenses", {
                        "expense_type": et, "amount": ea,
                        "description": edesc, "expense_date": str(ed)
                    })
                    st.rerun()
    with t2:
        exps = sorted(
            db_select("expenses"),
            key=lambda x: x.get('expense_date', ''), reverse=True)
        if exps:
            st.metric("Total",
                      f"₹ {sum(float(e.get('amount', 0)) for e in exps):,.2f}")
            st.dataframe(
                pd.DataFrame([{
                    "Date": e.get('expense_date', ''),
                    "Type": e.get('expense_type', ''),
                    "Amount": f"₹{float(e.get('amount', 0)):,.2f}",
                    "Desc": e.get('description', '')
                } for e in exps]),
                use_container_width=True, hide_index=True)


# ============================================================
# PAGE: REPORTS
# ============================================================
def page_reports():
    render_page_header("📊 Reports", "Financial reports & analysis")

    rc1, rc2, rc3 = st.columns(3)
    with rc1:
        period = st.selectbox(
            "Period", ["Daily", "Weekly", "Monthly", "Yearly", "Custom"])
    t = date.today()
    if period == "Custom":
        with rc2:
            sd = st.date_input("From", value=t - timedelta(30))
        with rc3:
            ed = st.date_input("To", value=t)
    else:
        sd, ed = get_period_dates(period)

    ptn = ["All"] + [p['name'] for p in db_select("pooja_types", "name")]
    pf = st.selectbox("Pooja Filter", ptn)
    bills = db_select("bills",
                      gte_filters={"bill_date": sd},
                      lte_filters={"bill_date": ed})
    exps = db_select("expenses",
                     gte_filters={"expense_date": sd},
                     lte_filters={"expense_date": ed})
    if pf != "All":
        bills = [b for b in bills if b.get('pooja_type') == pf]
    ti = sum(float(b.get('amount', 0)) for b in bills)
    te = sum(float(e.get('amount', 0)) for e in exps)

    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        st.markdown(
            f'<div class="metric-card income"><h3>Income</h3>'
            f'<h2>₹{ti:,.2f}</h2></div>', unsafe_allow_html=True)
    with mc2:
        st.markdown(
            f'<div class="metric-card expense"><h3>Expenses</h3>'
            f'<h2>₹{te:,.2f}</h2></div>', unsafe_allow_html=True)
    with mc3:
        st.markdown(
            f'<div class="metric-card balance"><h3>Balance</h3>'
            f'<h2>₹{ti - te:,.2f}</h2></div>', unsafe_allow_html=True)

    rt1, rt2, rt3 = st.tabs(["Income", "Expenses", "Charts"])
    with rt1:
        if bills:
            df = pd.DataFrame([{
                "Bill": b.get('bill_no', ''),
                "Date": b.get('bill_date', ''),
                "Pooja": b.get('pooja_type', ''),
                "Amount": float(b.get('amount', 0))
            } for b in bills])
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.download_button("📥 CSV", df.to_csv(index=False),
                               "income.csv")
    with rt2:
        if exps:
            st.dataframe(pd.DataFrame([{
                "Date": e.get('expense_date', ''),
                "Type": e.get('expense_type', ''),
                "Amount": float(e.get('amount', 0))
            } for e in exps]), use_container_width=True, hide_index=True)
    with rt3:
        if bills or exps:
            st.bar_chart(pd.DataFrame({
                "Cat": ["Income", "Expenses"],
                "₹": [ti, te]
            }).set_index("Cat"))


# ============================================================
# PAGE: ASSETS (Enhanced with Barcode Generation)
# ============================================================
def page_assets():
    render_page_header("🏷️ Assets", "Manage & Generate Barcodes")

    t1, t2, t3 = st.tabs(["➕ Add Asset", "📋 Asset List",
                            "🏷️ Barcode Generator"])

    with t1:
        with st.form("af", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                at = st.text_input("🏷️ Asset Tag *")
                an = st.text_input("📦 Asset Name *")
                sn = st.text_input("🔢 Serial Number")
                aq = st.number_input("📊 Quantity", min_value=1,
                                     value=1, step=1)
            with c2:
                dn = st.text_input("🙏 Donor Name")
                dd = st.date_input("📅 Donation Date",
                                   min_value=MIN_DATE, max_value=MAX_DATE)
                al = st.text_input("📍 Location")
                ai = st.file_uploader("📷 Image",
                                      type=['jpg', 'jpeg', 'png'])
            adesc = st.text_area("📝 Notes", height=60)
            auto_barcode = st.checkbox("🏷️ Auto-generate barcode", value=True)

            if st.form_submit_button("✅ Add Asset", use_container_width=True):
                if at.strip() and an.strip():
                    asset_data = {
                        "asset_tag": at.strip(),
                        "asset_name": an.strip(),
                        "serial_no": sn,
                        "donor_name": dn,
                        "donation_date": str(dd),
                        "image_url": file_to_base64(ai),
                        "description": adesc
                    }

                    # Generate and store barcode
                    if auto_barcode:
                        bc_img, bc_bytes = generate_barcode_image(at.strip())
                        if bc_img:
                            asset_data['barcode_url'] = bc_img

                    result = db_insert("assets", asset_data)
                    if result:
                        st.success(f"✅ Asset '{an}' added!")
                        if auto_barcode and bc_img:
                            st.markdown(
                                f'<div class="barcode-container">'
                                f'<p><b>Generated Barcode for: {at}</b></p>'
                                f'<img src="{bc_img}" style="max-width:300px">'
                                f'</div>', unsafe_allow_html=True)
                        st.rerun()

    with t2:
        # Search
        as1, as2 = st.columns(2)
        with as1:
            a_search = st.text_input("🔍 Search Assets", key="asset_search")
        with as2:
            a_sort = st.selectbox("Sort By",
                                  ["Name", "Tag", "Date"], key="asset_sort")

        assets = db_select("assets")

        if a_search:
            assets = [a for a in assets
                      if a_search.lower() in
                      f"{a.get('asset_tag', '')} {a.get('asset_name', '')} {a.get('donor_name', '')}".lower()]

        st.markdown(f"**Total Assets: {len(assets)}**")

        for a in assets:
            with st.expander(
                    f"🏷️ {a.get('asset_tag', '')} | "
                    f"{a.get('asset_name', '')}"):
                ac1, ac2, ac3 = st.columns([2, 2, 1])

                with ac1:
                    for l, k in [
                        ("Tag", "asset_tag"), ("Name", "asset_name"),
                        ("Serial", "serial_no"), ("Donor", "donor_name"),
                        ("Date", "donation_date"),
                        ("Notes", "description")
                    ]:
                        val = a.get(k, 'N/A')
                        if val and str(val) != 'None':
                            st.write(f"**{l}:** {val}")

                with ac2:
                    # Show barcode
                    if a.get('barcode_url'):
                        st.markdown(
                            f'<div class="barcode-container">'
                            f'<img src="{a["barcode_url"]}" '
                            f'style="max-width:250px">'
                            f'</div>', unsafe_allow_html=True)
                    else:
                        # Generate on-the-fly
                        bc_img, bc_bytes = generate_barcode_image(
                            a.get('asset_tag', ''))
                        if bc_img:
                            st.markdown(
                                f'<div class="barcode-container">'
                                f'<img src="{bc_img}" '
                                f'style="max-width:250px">'
                                f'</div>', unsafe_allow_html=True)

                    if (a.get('image_url')
                            and a['image_url'].startswith('data:')):
                        st.markdown(
                            f'<img src="{a["image_url"]}" width="130" '
                            f'style="border-radius:10px">',
                            unsafe_allow_html=True)

                with ac3:
                    # Generate barcode for this asset
                    tag = a.get('asset_tag', '')
                    bc_img_dl, bc_bytes_dl = generate_barcode_image(tag)

                    if bc_bytes_dl:
                        st.download_button(
                            "📥 Barcode PNG",
                            data=bc_bytes_dl,
                            file_name=f"barcode_{tag}.png",
                            mime="image/png",
                            key=f"bc_png_{a['id']}")

                    if PDF_AVAILABLE:
                        bc_pdf = generate_asset_barcode_pdf(
                            tag, a.get('asset_name', ''), bc_bytes_dl)
                        if bc_pdf:
                            st.download_button(
                                "📥 Label PDF",
                                data=bc_pdf,
                                file_name=f"label_{tag}.pdf",
                                mime="application/pdf",
                                key=f"bc_pdf_{a['id']}")

                    if st.button("🏷️ Regenerate",
                                 key=f"regen_{a['id']}"):
                        new_bc, _ = generate_barcode_image(tag)
                        if new_bc:
                            db_update("assets",
                                      {"barcode_url": new_bc},
                                      "id", a['id'])
                            st.success("✅ Barcode regenerated!")
                            st.rerun()

                    if st.button("🗑️ Delete",
                                 key=f"da_{a['id']}"):
                        db_delete("assets", "id", a['id'])
                        st.rerun()

    with t3:
        st.markdown("### 🏷️ Bulk Barcode Generator")
        st.markdown("""
        Generate barcodes for all assets or specific ones.
        Barcodes can be printed as labels for physical asset tracking.
        """)

        gen_mode = st.radio(
            "Generate for",
            ["All Assets", "Specific Asset", "Custom Text"],
            horizontal=True)

        if gen_mode == "All Assets":
            all_assets = db_select("assets", "id,asset_tag,asset_name")
            if all_assets:
                st.write(f"**{len(all_assets)} assets found**")
                if st.button("🏷️ Generate All Barcodes",
                             use_container_width=True, type="primary"):
                    with st.spinner("Generating barcodes..."):
                        cols_per_row = 3
                        for i in range(0, len(all_assets), cols_per_row):
                            cols = st.columns(cols_per_row)
                            for j, asset in enumerate(
                                    all_assets[i:i + cols_per_row]):
                                with cols[j]:
                                    bc_img, bc_bytes = generate_barcode_image(
                                        asset['asset_tag'])
                                    st.markdown(
                                        f'<div class="barcode-container">'
                                        f'<p><b>{asset["asset_name"]}</b></p>'
                                        f'<img src="{bc_img}" '
                                        f'style="max-width:200px">'
                                        f'<p style="font-size:0.8em">'
                                        f'{asset["asset_tag"]}</p>'
                                        f'</div>', unsafe_allow_html=True)

                                    # Update DB with barcode
                                    db_update("assets",
                                              {"barcode_url": bc_img},
                                              "id", asset['id'])

                    st.success("✅ All barcodes generated!")

        elif gen_mode == "Specific Asset":
            assets_list = db_select("assets", "id,asset_tag,asset_name")
            if assets_list:
                asset_opts = {
                    f"{a['asset_tag']} - {a['asset_name']}": a
                    for a in assets_list}
                selected = st.selectbox("Select Asset",
                                        list(asset_opts.keys()))
                if selected and st.button("🏷️ Generate",
                                          type="primary"):
                    asset = asset_opts[selected]
                    bc_img, bc_bytes = generate_barcode_image(
                        asset['asset_tag'])
                    st.markdown(
                        f'<div class="barcode-container">'
                        f'<h3>{asset["asset_name"]}</h3>'
                        f'<img src="{bc_img}" style="max-width:350px">'
                        f'<p>{asset["asset_tag"]}</p>'
                        f'</div>', unsafe_allow_html=True)

                    col_a, col_b = st.columns(2)
                    with col_a:
                        if bc_bytes:
                            st.download_button(
                                "📥 Download PNG", data=bc_bytes,
                                file_name=f"barcode_{asset['asset_tag']}.png",
                                mime="image/png")
                    with col_b:
                        if PDF_AVAILABLE:
                            bc_pdf = generate_asset_barcode_pdf(
                                asset['asset_tag'],
                                asset['asset_name'], bc_bytes)
                            if bc_pdf:
                                st.download_button(
                                    "📥 Download Label PDF",
                                    data=bc_pdf,
                                    file_name=f"label_{asset['asset_tag']}.pdf",
                                    mime="application/pdf")

        else:  # Custom Text
            custom_text = st.text_input("Enter text for barcode")
            custom_name = st.text_input("Label name (optional)")
            if custom_text and st.button("🏷️ Generate",
                                         type="primary"):
                bc_img, bc_bytes = generate_barcode_image(custom_text)
                st.markdown(
                    f'<div class="barcode-container">'
                    f'{"<h3>" + custom_name + "</h3>" if custom_name else ""}'
                    f'<img src="{bc_img}" style="max-width:350px">'
                    f'<p>{custom_text}</p>'
                    f'</div>', unsafe_allow_html=True)

                if bc_bytes:
                    st.download_button(
                        "📥 Download PNG", data=bc_bytes,
                        file_name=f"barcode_{custom_text}.png",
                        mime="image/png")


# ============================================================
# PAGE: SETTINGS (Enhanced with Amman Image Upload)
# ============================================================
def page_settings():
    render_page_header("⚙️ Settings", "Temple Configuration")

    t1, t2, t3, t4 = st.tabs(
        ["🖼️ Amman Image", "🙏 Pooja Types",
         "💸 Expense Types", "📢 News Ticker"])

    # ---- TAB 1: AMMAN IMAGE UPLOAD ----
    with t1:
        st.markdown("### 🖼️ Amman Image Management")
        st.markdown("""
        Upload the Amman image to display on:
        - 🔐 **Login page** (round circle with glow)
        - 📑 **Every page banner** (left & right)
        - 🧾 **Bill PDFs** (logo on bills)
        - 📱 **Sidebar** (small round icon)
        - 🎨 **Login background** (faded watermark)
        """)

        current_img = get_amman_image()

        # Show current image
        st.markdown("#### Current Amman Image")
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown(f"""
            <div style="text-align:center;padding:20px;
                background:linear-gradient(135deg,#fff5ee,#ffe4c4);
                border-radius:15px;border:2px solid #ff6b35;">
                <img src="{current_img}" style="width:150px;height:150px;
                    border-radius:50%;border:4px solid #ff6b35;
                    object-fit:cover;
                    box-shadow:0 0 20px rgba(255,107,53,0.4);">
                <p style="margin-top:10px;color:#8B0000;font-weight:bold;">
                    Current Image</p>
            </div>
            """, unsafe_allow_html=True)

        with c2:
            st.markdown("#### Upload New Amman Image")
            st.info(
                "📌 **Recommended:** Square image (500x500px or larger), "
                "JPG or PNG format. The image will appear in round circles.")

            new_photo = st.file_uploader(
                "Choose Amman image (JPG/PNG)",
                type=['jpg', 'jpeg', 'png'],
                key="amman_upload_settings",
                help="Upload a high-quality photo of Amman deity")

            if new_photo:
                # Preview
                img_b64 = file_to_base64(new_photo)
                st.markdown("**Preview:**")
                pc1, pc2, pc3 = st.columns(3)
                with pc1:
                    st.markdown(f"""
                    <div style="text-align:center;">
                        <p>Login (Large)</p>
                        <img src="{img_b64}" style="width:120px;height:120px;
                            border-radius:50%;border:4px solid #ff6b35;
                            object-fit:cover;">
                    </div>
                    """, unsafe_allow_html=True)
                with pc2:
                    st.markdown(f"""
                    <div style="text-align:center;">
                        <p>Banner (Small)</p>
                        <img src="{img_b64}" style="width:50px;height:50px;
                            border-radius:50%;border:2px solid #ff6b35;
                            object-fit:cover;">
                    </div>
                    """, unsafe_allow_html=True)
                with pc3:
                    st.markdown(f"""
                    <div style="text-align:center;">
                        <p>Sidebar</p>
                        <img src="{img_b64}" style="width:70px;height:70px;
                            border-radius:50%;border:3px solid #ff6b35;
                            object-fit:cover;">
                    </div>
                    """, unsafe_allow_html=True)

                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("✅ Save Amman Image",
                                 use_container_width=True, type="primary"):
                        save_amman_image_to_db(img_b64)
                        st.success("✅ Amman image updated successfully!")
                        st.balloons()
                        time.sleep(1)
                        st.rerun()
                with bc2:
                    if st.button("❌ Cancel", use_container_width=True):
                        st.rerun()

            st.markdown("---")

            # Reset option
            if st.session_state.get('custom_amman_photo') or True:
                if st.button("🔄 Reset to Default Image",
                             use_container_width=True):
                    st.session_state['custom_amman_photo'] = None
                    if DB_CONNECTED:
                        try:
                            existing = db_select("temple_settings",
                                                 filters={"key": "amman_image"})
                            if existing:
                                db_delete("temple_settings",
                                          "key", "amman_image")
                        except:
                            pass
                    st.success("✅ Reset to default!")
                    st.rerun()

    # ---- TAB 2: POOJA TYPES ----
    with t2:
        st.markdown("### 🙏 Pooja Types")
        for p in db_select("pooja_types"):
            c1, c2 = st.columns([5, 1])
            with c1:
                st.write(f"🙏 **{p['name']}** — ₹{p.get('amount', 0)}")
            with c2:
                if st.button("🗑️", key=f"dp_{p['id']}"):
                    db_delete("pooja_types", "id", p['id'])
                    st.rerun()
        with st.form("apt", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nn = st.text_input("Pooja Name")
            with c2:
                na = st.number_input("Amount", min_value=0.0, step=10.0)
            if st.form_submit_button("➕ Add"):
                if nn.strip():
                    db_insert("pooja_types",
                              {"name": nn.strip(), "amount": na})
                    st.rerun()

    # ---- TAB 3: EXPENSE TYPES ----
    with t3:
        st.markdown("### 💸 Expense Types")
        for e in db_select("expense_types"):
            c1, c2 = st.columns([5, 1])
            with c1:
                st.write(f"💸 **{e['name']}**")
            with c2:
                if st.button("🗑️", key=f"de_{e['id']}"):
                    db_delete("expense_types", "id", e['id'])
                    st.rerun()
        with st.form("aet", clear_on_submit=True):
            nn = st.text_input("Expense Type Name")
            if st.form_submit_button("➕ Add"):
                if nn.strip():
                    db_insert("expense_types", {"name": nn.strip()})
                    st.rerun()

    # ---- TAB 4: NEWS TICKER ----
    with t4:
        st.markdown("### 📢 News Ticker")
        for n in db_select("news_ticker"):
            c1, c2, c3 = st.columns([4, 1, 1])
            with c1:
                icon = '🟢' if n.get('is_active') else '🔴'
                st.write(f"{icon} {n['message']}")
            with c2:
                if st.button("Toggle", key=f"tn_{n['id']}"):
                    db_update("news_ticker",
                              {"is_active": not n.get('is_active', True)},
                              "id", n['id'])
                    st.rerun()
            with c3:
                if st.button("🗑️", key=f"dn_{n['id']}"):
                    db_delete("news_ticker", "id", n['id'])
                    st.rerun()
        with st.form("an", clear_on_submit=True):
            nm = st.text_input("News Message")
            if st.form_submit_button("➕ Add"):
                if nm.strip():
                    db_insert("news_ticker",
                              {"message": nm.strip(), "is_active": True})
                    st.rerun()


# ============================================================
# PAGE: USERS
# ============================================================
def page_users():
    render_page_header("👥 Users", "User Management")

    if st.session_state.user_role != 'admin':
        st.error("Admin only!")
        return

    t1, t2 = st.tabs(["➕ Create", "📋 List"])
    with t1:
        with st.form("cu", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nu = st.text_input("Username")
                np_ = st.text_input("Password", type="password")
            with c2:
                cp = st.text_input("Confirm Password", type="password")
                nr = st.selectbox("Role", ["user", "admin"])
            if st.form_submit_button("➕ Create User",
                                     use_container_width=True):
                if (nu and np_ and np_ == cp
                        and not db_select("users",
                                          filters={"username": nu})):
                    db_insert("users", {
                        "username": nu, "password_hash": np_, "role": nr
                    })
                    st.success(f"✅ User '{nu}' created!")
                    st.rerun()
                elif np_ != cp:
                    st.error("Passwords don't match!")
    with t2:
        for u in db_select("users"):
            c1, c2 = st.columns([5, 1])
            with c1:
                icon = '👑' if u.get('role') == 'admin' else '👤'
                st.write(f"{icon} **{u['username']}** ({u.get('role', '')})")
            with c2:
                if u['username'] != 'admin':
                    if st.button("🗑️", key=f"du_{u['id']}"):
                        db_delete("users", "id", u['id'])
                        st.rerun()


# ============================================================
# PAGE: SAMAYA VAKUPPU
# ============================================================
def page_samaya():
    render_page_header("📚 Samaya Vakuppu", "Student Management")

    t1, t2 = st.tabs(["➕ Add", "📋 List"])
    with t1:
        with st.form("sv", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                sn = st.text_input("Name *")
                sd = st.date_input("DOB", value=date(2010, 1, 1),
                                   min_value=MIN_DATE, max_value=MAX_DATE)
                sa = st.text_area("Address", height=60)
                spt = st.selectbox("Parent Type", ["Father", "Mother"])
                spn = st.text_input("Parent Name")
            with c2:
                sbd = st.date_input("Bond Date",
                                    min_value=MIN_DATE, max_value=MAX_DATE)
                sbk = st.text_input("Bank")
                sbr = st.text_input("Branch")
                sbn = st.text_input("Bond No")
                sbf = st.file_uploader(
                    "Bond", type=['jpg', 'jpeg', 'png', 'pdf'], key="svb")
                sph = st.file_uploader(
                    "Photo", type=['jpg', 'jpeg', 'png'], key="svp")
            if st.form_submit_button("✅", use_container_width=True):
                if sn.strip():
                    db_insert("samaya_vakuppu", {
                        "student_name": sn.strip(), "dob": str(sd),
                        "address": sa, "parent_name": spn,
                        "parent_type": spt,
                        "bond_issue_date": str(sbd),
                        "scanned_bond_url": file_to_base64(sbf),
                        "photo_url": file_to_base64(sph),
                        "bond_issuing_bank": sbk,
                        "branch_of_bank": sbr, "bond_no": sbn
                    })
                    st.rerun()
    with t2:
        for s in db_select("samaya_vakuppu"):
            with st.expander(f"👤 {s['student_name']}"):
                for l, k in [
                    ("Name", "student_name"), ("DOB", "dob"),
                    ("Address", "address"), ("Parent", "parent_name"),
                    ("Bond", "bond_no")
                ]:
                    st.write(f"**{l}:** {s.get(k, 'N/A')}")
                if st.button("🗑️", key=f"ds_{s['id']}"):
                    db_delete("samaya_vakuppu", "id", s['id'])
                    st.rerun()


# ============================================================
# PAGE: THIRUMANA MANDAPAM
# ============================================================
def page_thirumana():
    render_page_header("💒 Thirumana Mandapam", "Bond Management")

    t1, t2 = st.tabs(["➕ Add", "📋 List"])
    with t1:
        with st.form("tm", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                tn = st.text_input("Name *")
                ta = st.text_area("Address", height=60)
                tb = st.text_input("Bond No")
                td = st.date_input("Date",
                                   min_value=MIN_DATE, max_value=MAX_DATE)
            with c2:
                tam = st.number_input("Amount", min_value=0.0, step=100.0)
                tnb = st.number_input("No of Bonds", min_value=0, step=1)
                ts = st.file_uploader(
                    "Scan", type=['jpg', 'jpeg', 'png', 'pdf'], key="tms")
                tp = st.file_uploader(
                    "Photo", type=['jpg', 'jpeg', 'png'], key="tmp")
            if st.form_submit_button("✅", use_container_width=True):
                if tn.strip():
                    db_insert("thirumana_mandapam", {
                        "name": tn.strip(), "address": ta,
                        "bond_no": tb, "bond_issued_date": str(td),
                        "amount": tam, "no_of_bonds": tnb,
                        "scan_copy_url": file_to_base64(ts),
                        "photo_url": file_to_base64(tp)
                    })
                    st.rerun()
    with t2:
        for r in db_select("thirumana_mandapam"):
            with st.expander(
                    f"👤 {r['name']} | ₹{r.get('amount', 0)}"):
                for l, k in [
                    ("Name", "name"), ("Address", "address"),
                    ("Bond", "bond_no"), ("Date", "bond_issued_date"),
                    ("Amount", "amount")
                ]:
                    st.write(f"**{l}:** {r.get(k, 'N/A')}")
                if st.button("🗑️", key=f"dt_{r['id']}"):
                    db_delete("thirumana_mandapam", "id", r['id'])
                    st.rerun()


# ============================================================
# SIDEBAR
# ============================================================
def render_sidebar():
    with st.sidebar:
        amman_img = get_amman_image()
        st.markdown(f"""
        <div class="sidebar-amman">
            <img src="{amman_img}" alt="Amman">
        </div>
        <div style="text-align:center;padding:5px;
            background:linear-gradient(135deg,#ff6b35,#f7c948);
            border-radius:8px;margin-bottom:10px;">
            <p style="color:#5a1a00;margin:0;font-weight:600;
                font-size:0.7em;">
                Sree Bhadreshwari Amman<br>Temple Management</p>
        </div>
        <div style="color:#ccc;padding:3px 10px;font-size:0.8em;">
            👤 <b style="color:#f7c948">{st.session_state.username}</b>
            ({st.session_state.user_role})
        </div>
        """, unsafe_allow_html=True)
        st.markdown("---")

        pages = [
            ("🏠 Dashboard", "Dashboard"),
            ("👥 Devotees", "Devotees"),
            ("🧾 Billing", "Billing"),
            ("💸 Expenses", "Expenses"),
            ("📊 Reports", "Reports"),
            ("🏷️ Assets", "Assets"),
            ("📚 Samaya Vakuppu", "Samaya"),
            ("💒 Thirumana", "Thirumana"),
            ("⚙️ Settings", "Settings"),
            ("👥 Users", "Users"),
        ]
        for l, p in pages:
            if p == "Users" and st.session_state.user_role != 'admin':
                continue
            if st.button(l, key=f"n_{p}", use_container_width=True):
                st.session_state.current_page = p
                st.rerun()

        st.markdown("---")
        if st.button("🚪 Logout", key="lo", use_container_width=True):
            for k in ['logged_in', 'username', 'user_role', 'current_page']:
                st.session_state[k] = defaults[k]
            st.rerun()

        st.markdown(
            '<div style="text-align:center;padding:15px 0;color:#555;'
            'font-size:0.65em;">v3.0 🙏 அம்மே நாராயணா 🙏</div>',
            unsafe_allow_html=True)


# ============================================================
# MAIN
# ============================================================
def main():
    # Apply CSS globally
    amman_img = get_amman_image()
    st.markdown(get_custom_css(amman_img), unsafe_allow_html=True)

    if not st.session_state.logged_in:
        page_login()
    else:
        render_sidebar()
        pm = {
            "Dashboard": page_dashboard,
            "Devotees": page_devotee_enrollment,
            "Billing": page_billing,
            "Expenses": page_expenses,
            "Reports": page_reports,
            "Assets": page_assets,
            "Samaya": page_samaya,
            "Thirumana": page_thirumana,
            "Settings": page_settings,
            "Users": page_users,
        }
        pm.get(st.session_state.current_page, page_dashboard)()


if __name__ == "__main__":
    main()
